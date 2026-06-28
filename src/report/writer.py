"""Populate the report template (report_template.xlsx) with computed figures.

Each row is matched by (Section, Metric) and filled with Value, Limit,
Utilization, Status, and the traceability cell: graph path -> source doc/page.
Only engine output is written; the narrative is added as a separate sheet so the
numeric report and the prose are physically separated.
"""
from __future__ import annotations

import openpyxl


def write_report(template_path: str, out_path: str, figures: list[dict],
                 narrative: str) -> None:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    by_metric = {f["metric"]: f for f in figures}
    header = [c.value for c in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header)}
    src_col = None
    for i, name in enumerate(header):
        if name and str(name).startswith("Source"):
            src_col = i + 1

    for row in range(2, ws.max_row + 1):
        metric = ws.cell(row=row, column=col["Metric"]).value
        fig = by_metric.get(metric)
        if not fig:
            continue
        if fig.get("status") == "ERROR":
            ws.cell(row=row, column=col["Status"]).value = "ERROR"
            if src_col:
                ws.cell(row=row, column=src_col).value = fig.get("error", "")
            continue
        ws.cell(row=row, column=col["Value"]).value = fig["value"]
        ws.cell(row=row, column=col["Limit"]).value = fig["limit"]
        ws.cell(row=row, column=col["Utilization"]).value = fig["utilization"]
        ws.cell(row=row, column=col["Status"]).value = fig["status"]
        if src_col:
            cite = fig["citation"]
            ws.cell(row=row, column=src_col).value = (
                f"{fig['graph_path']}  →  {cite['source_doc']} "
                f"p.{cite['page']} ({cite['chunk_id']})"
            )

    # Narrative on a separate sheet — prose is never mixed with the figures.
    nsheet = wb.create_sheet("Narrative")
    nsheet["A1"] = "LLM narrative (commentary only — firewall-verified, no new numbers)"
    nsheet["A2"] = narrative

    wb.save(out_path)
